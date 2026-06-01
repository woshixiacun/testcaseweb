"""把选中的 case 按版本分 sheet 导出为 xlsx，与旧 Node 实现行为对齐。

- 每个版本一个 sheet；sheet 名按 _versions.json 顺序，未知/无版本排在后面
- 一行一个 case；Preconditions + Test Steps 合并到最后一个单元格
- 表头加粗 + 浅紫底；数据顶端对齐 + 自动换行
- 文件名带版本标签；多版本拼接，非法字符替换为下划线
"""
from __future__ import annotations

import io
import re
import time
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 与前端 StatusSelect 一致的可读标签
STATUS_LABEL = {
    "pending": "Not Run",
    "success": "Passed",
    "fail": "Failed",
    "blocked": "Blocked",
    "": "Not Run",
}
TYPE_LABEL = {
    "uncategorized": "Uncategorized",
    "manual": "Manual",
    "auto": "Automated",
    "": "Uncategorized",
}

NO_VERSION = "(No Version)"


def _iterations_value(val: Any) -> Any:
    """迭代次数：能转成整数就写数字，否则原样透传字符串（空值写空串）。"""
    if val is None or val == "":
        return ""
    try:
        return int(val)
    except (TypeError, ValueError):
        return str(val)

COLUMNS = [
    ("Case Name", 36),
    ("Case ID", 16),
    ("Version", 22),
    ("Requirement", 16),
    ("Case Status", 12),
    ("Case Type", 14),
    ("Number of Iteration", 16),
    ("Preconditions & Test Steps", 60),
]


def _merge_steps_cell(c: dict) -> str:
    """把一个 case 的 Preconditions + Test Steps 合并成单元格文本。"""
    lines: list[str] = []
    pre = c.get("precondition") or ""
    lines.append(f"Preconditions: {pre if pre else '(none)'}")
    lines.append("")
    lines.append("Test Steps:")
    steps = c.get("steps") if isinstance(c.get("steps"), list) else []
    if not steps:
        lines.append("(none)")
    else:
        for i, s in enumerate(steps, start=1):
            action = (s.get("operation") or "").strip()
            expected = (s.get("expected") or "").strip()
            lines.append(f"{i}. {action}；{expected}")
    return "\n".join(lines)


_INVALID_SHEET_CHARS = re.compile(r"[\\/?*\[\]:]")


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """sheet 名清洗：去非法字符，截断 31 字符，保证非空且唯一。"""
    n = _INVALID_SHEET_CHARS.sub(" ", str(name or "")).strip()[:31]
    if not n:
        n = "Sheet"
    base = n
    i = 2
    while n.lower() in used:
        suffix = f" ({i})"
        n = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(n.lower())
    return n


def build_export_xlsx(
    cases: list[dict], versions: list[dict]
) -> tuple[bytes, str]:
    """构建多 sheet xlsx，返回 (bytes, filename)。"""
    # 按版本分组
    order = [v.get("edition") for v in versions]
    groups: dict[str, list[dict]] = {}
    for c in cases:
        key = c.get("version") or NO_VERSION
        groups.setdefault(key, []).append(c)

    def sort_key(k: str) -> tuple[int, str]:
        try:
            return (order.index(k), "")
        except ValueError:
            return (10**9, k)

    sorted_keys = sorted(groups.keys(), key=sort_key)

    wb = Workbook()
    # 删除默认 sheet，由我们按需创建
    wb.remove(wb.active)

    header_font = Font(bold=True, name="Calibri", size=11)
    header_fill = PatternFill("solid", fgColor="EFEAFB")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="top", wrap_text=True)

    used_names: set[str] = set()
    for key in sorted_keys:
        ws = wb.create_sheet(title=_safe_sheet_name(key, used_names))

        # 表头
        for ci, (header, width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=ci, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            ws.column_dimensions[get_column_letter(ci)].width = width

        # 数据
        for ri, c in enumerate(groups[key], start=2):
            row_values = [
                c.get("caseName", "") or "",
                c.get("caseId", "") or "",
                c.get("version", "") or "",
                c.get("requirementDir", "") or "",
                STATUS_LABEL.get(c.get("caseStatus"), STATUS_LABEL[""]),
                TYPE_LABEL.get(c.get("caseType"), TYPE_LABEL[""]),
                _iterations_value(c.get("iterations")),
                _merge_steps_cell(c),
            ]
            for ci, val in enumerate(row_values, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = body_align

    # 文件名：单版本用其名字，多版本拼接，非法字符替换
    real_versions = [k for k in sorted_keys if k != NO_VERSION]
    label = "_".join(real_versions) if real_versions else "testcases"
    label = re.sub(r"[^a-zA-Z0-9_\-.]", "_", label)[:80] or "testcases"
    filename = f"{label}_{int(time.time() * 1000)}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), filename
