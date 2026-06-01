"""导入 xlsx：把人工写的、行/列对不齐的测试用例 Excel 解析成标准 case。

支持的列（表头大小写不敏感、空格忽略、模糊别名）：
  Area / ID / Test Case / Iteration / Test Procedure / Check Point / Type / Requirement doc

关键规则：
- 单值字段（Area / ID / Test Case / Type / Iteration）：可能写在任意一行，或者是合并单元格。
  做法：先把合并单元格的左上角值"展开"到整个 merge 范围；再按 ID 列**向下**前向填充
  （每个空 ID 行继承下一个非空 ID 行的值），由此把同一 case 的连续行聚合成一组。
- 多行字段（Test Procedure / Check Point / Requirement doc）：每行一个值。
  Test Procedure 用 Pre/Prerequisite/前提/Steps/步骤 等关键字切分前置条件与步骤；
  没有关键字就全当 step。Check Point 一行 = 一个 expected。
  Steps 与 Check Points 行数可能不齐，按 max 对齐填充。
"""
from __future__ import annotations

import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

# 列名别名 → 标准键
COL_ALIASES = {
    "area": "area",
    "id": "id",
    "caseid": "id",
    "case id": "id",
    "testcase": "case_name",
    "test case": "case_name",
    "casename": "case_name",
    "case name": "case_name",
    "iteration": "iteration",
    "iterations": "iteration",
    "number of iteration": "iteration",
    "testprocedure": "procedure",
    "test procedure": "procedure",
    "procedure": "procedure",
    "steps": "procedure",
    "step": "procedure",
    "checkpoint": "checkpoint",
    "check point": "checkpoint",
    "expected": "checkpoint",
    "expected result": "checkpoint",
    "type": "type",
    "case type": "type",
    "requirement": "requirement",
    "requirementdoc": "requirement",
    "requirement doc": "requirement",
    "requirements": "requirement",
}

# Test Procedure 单元格里 Pre/Steps 的标记前缀（行起始即视为标记，不区分大小写）
_PRE_MARK = re.compile(r"^\s*(pre[a-z]*|prerequisites?|前提)\s*[:：]?\s*$", re.I)
_STEP_MARK = re.compile(r"^\s*(steps?|步骤)\s*[:：]?\s*$", re.I)
# 行首的列表序号："1." / "2)" / "(3)" / "1、" 等
_NUM_PREFIX = re.compile(r"^\s*[\(（]?\s*\d+\s*[\)）\.\、:：]\s*")


def _norm_header(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _strip_numbering(line: str) -> str:
    return _NUM_PREFIX.sub("", line).strip()


def _map_type(raw: Any) -> str:
    s = str(raw or "").strip().lower()
    if s in ("auto", "automated", "automation", "自动", "自动化"):
        return "auto"
    if s in ("manual", "手动", "手工"):
        return "manual"
    return "uncategorized"


def _split_lines(val: Any) -> list[str]:
    """把单元格值切成行；非字符串原样转字符串；保留空白裁剪后的非空行。"""
    if val is None:
        return []
    text = str(val).replace("\r\n", "\n").replace("\r", "\n")
    return [line.rstrip() for line in text.split("\n")]


def _parse_procedure_lines(lines: list[str]) -> tuple[list[str], list[str]]:
    """把 Test Procedure 的多行内容切成 (前置条件行, 步骤行)。

    规则：
    - 出现 Steps:/步骤: 标记之后，所有后续行算 step；之前所有行算 pre。
    - 出现 Pre:/Prerequisite: 标记后到 Steps 标记之间算 pre。
    - 标记行本身丢弃。
    - 没有任何 Steps 标记 → 全部算 step，pre 为空（与你的描述一致：默认都是 steps）。
    - 数字编号（"1. xxx"）一律去掉。
    """
    pre: list[str] = []
    steps: list[str] = []
    in_steps = False
    saw_step_mark = False
    # 第一遍：扫一下有没有 Steps 标记，决定无标记时全归 steps
    for ln in lines:
        if _STEP_MARK.match(ln):
            saw_step_mark = True
            break

    in_pre_section = False
    for ln in lines:
        if not ln.strip():
            continue
        if _STEP_MARK.match(ln):
            in_steps = True
            in_pre_section = False
            continue
        if _PRE_MARK.match(ln):
            in_steps = False
            in_pre_section = True
            continue
        cleaned = _strip_numbering(ln)
        if not cleaned:
            continue
        if in_steps:
            steps.append(cleaned)
        elif in_pre_section or saw_step_mark:
            # 有 Steps 标记 → 标记前的内容算 pre（不论是否有 Pre 标记）
            pre.append(cleaned)
        else:
            # 没有任何 Steps 标记 → 默认全是 steps
            steps.append(cleaned)
    return pre, steps


def _expand_merged(ws) -> dict[tuple[int, int], Any]:
    """把合并单元格的左上角值展开到整个范围。返回 {(row,col): value}。"""
    grid: dict[tuple[int, int], Any] = {}
    # 先放入所有非合并单元格的值
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                grid[(cell.row, cell.column)] = cell.value
    # 合并单元格：用左上角的值填满整个矩形
    for mr in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(mr))
        anchor = ws.cell(row=min_row, column=min_col).value
        if anchor is None:
            continue
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                grid.setdefault((r, c), anchor)
    return grid


def _read_cell(grid: dict, r: int, c: int) -> Any:
    return grid.get((r, c))


def parse_xlsx(content: bytes) -> dict:
    """解析上传的 xlsx 字节，返回 {cases:[...], errors:[...]}。

    每个 case 的形状（与前端 case 对齐，但 caseStatus 固定 pending）：
      {area, caseId, caseName, caseType, iterations, requirementDir, precondition, steps[]}
    """
    import io

    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    grid = _expand_merged(ws)
    if ws.max_row < 2:
        return {"cases": [], "errors": ["Empty sheet"]}

    # --- 1) 解析表头：找出每个标准列在第几列 ---
    header_row = 1
    col_index: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = COL_ALIASES.get(_norm_header(_read_cell(grid, header_row, c)))
        if key and key not in col_index:
            col_index[key] = c

    missing = [k for k in ("id", "case_name") if k not in col_index]
    if missing:
        return {
            "cases": [],
            "errors": [
                f"Missing required column(s): {', '.join(missing)}. "
                f"Recognized headers: {sorted(set(COL_ALIASES))}"
            ],
        }

    id_col = col_index["id"]
    name_col = col_index["case_name"]
    area_col = col_index.get("area")
    type_col = col_index.get("type")
    iter_col = col_index.get("iteration")
    proc_col = col_index.get("procedure")
    chk_col = col_index.get("checkpoint")
    req_col = col_index.get("requirement")

    # --- 2) ID 列后向填充：每个空 ID 行继承下一个非空 ID 行的值，把同 case 的多行聚合 ---
    last_data_row = ws.max_row
    next_id_for: dict[int, Any] = {}
    cur: Any = None
    for r in range(last_data_row, header_row, -1):
        v = _read_cell(grid, r, id_col)
        if v is not None and str(v).strip() != "":
            cur = v
        next_id_for[r] = cur

    # --- 3) 按 ID 分组（连续相同 ID 的行算一个 case） ---
    groups: list[dict] = []
    cur_group: dict | None = None
    for r in range(header_row + 1, last_data_row + 1):
        cid = next_id_for.get(r)
        if cid is None or str(cid).strip() == "":
            continue
        cid_str = str(cid).strip()
        if cur_group is None or cur_group["id"] != cid_str:
            cur_group = {"id": cid_str, "rows": []}
            groups.append(cur_group)
        cur_group["rows"].append(r)

    # --- 4) 把每个 group 转成 case ---
    cases: list[dict] = []
    errors: list[str] = []

    def first_nonempty(rows: list[int], col: int | None) -> Any:
        if not col:
            return None
        for r in rows:
            v = _read_cell(grid, r, col)
            if v is not None and str(v).strip() != "":
                return v
        return None

    def collect_lines(rows: list[int], col: int | None) -> list[str]:
        """收集每行该列的内容；多行单元格内部用换行再切。空行跳过。"""
        if not col:
            return []
        out: list[str] = []
        for r in rows:
            v = _read_cell(grid, r, col)
            for line in _split_lines(v):
                if line.strip():
                    out.append(line)
        return out

    for g in groups:
        rows = g["rows"]
        case_name = first_nonempty(rows, name_col)
        if not case_name or not str(case_name).strip():
            errors.append(f"Row {rows[0]}: case '{g['id']}' has no Test Case name; skipped")
            continue

        area = str(first_nonempty(rows, area_col) or "").strip()
        type_raw = first_nonempty(rows, type_col)
        iter_raw = first_nonempty(rows, iter_col)

        # Test Procedure：跨多行汇总后再切 pre / steps
        proc_lines = collect_lines(rows, proc_col)
        pre_lines, step_lines = _parse_procedure_lines(proc_lines)

        # Check Point：每行一个 expected
        check_lines = [_strip_numbering(x) for x in collect_lines(rows, chk_col)]

        # Steps + Check Points 按 max 对齐
        n = max(len(step_lines), len(check_lines))
        steps_out: list[dict] = []
        for i in range(n):
            steps_out.append(
                {
                    "operation": step_lines[i] if i < len(step_lines) else "",
                    "expected": check_lines[i] if i < len(check_lines) else "",
                    "actualResult": "pending",
                    "actualNote": "",
                }
            )
        if not steps_out:
            steps_out = [
                {"operation": "", "expected": "", "actualResult": "pending", "actualNote": ""}
            ]

        # Requirement doc：每行一个值，逗号拼接（与多选下拉一致）
        req_lines = []
        seen = set()
        for ln in collect_lines(rows, req_col):
            v = ln.strip()
            if v and v not in seen:
                req_lines.append(v)
                seen.add(v)

        # Iteration：保留原样字符串（前端 iter-input 用的是 type=number，能识别就识别）
        iter_val = ""
        if iter_raw is not None:
            iter_str = str(iter_raw).strip()
            iter_val = iter_str

        # caseId 做基本清洗
        case_id = str(g["id"]).strip()

        cases.append(
            {
                "area": area,  # 由后端用作 tree 一级目录匹配
                "caseId": case_id,
                "caseName": str(case_name).strip(),
                "caseType": _map_type(type_raw),
                "caseStatus": "pending",  # 导入一律 Not Run
                "iterations": iter_val,
                "requirementDir": ", ".join(req_lines),
                "precondition": "\n".join(pre_lines) if pre_lines else "",
                "steps": steps_out,
            }
        )

    return {"cases": cases, "errors": errors}
